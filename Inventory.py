# Import Statements
import tkinter.messagebox
from ttkbootstrap import *
import openpyxl as opx
import sqlite3

# Window
root = Window(themename = "vapor")
root.geometry("800x500")
root.title("IM Inventory")

# Open inventory sheet
Custum = opx.load_workbook("Inventory.xlsx")
Sheeter = Custum.active

# Stringvar
Bile = StringVar()

# Style
my = Style()
my.configure("default.TButton",
             font=("Trebuchet MS",
                   18))

# Entry
E1 = Entry(root,
           font=("Trebuchet MS",
                 10),
           width = 30)
E1.place(relx = 0.5,
         rely = 0.3,
         anchor = "center")

# Label
L1 = Label(root,
           textvariable=Bile,
           font=("Comic Sans MS",
                 18),
           bootstyle="light")
L1.place(relx = 0.5,
         rely = 0.7,
         anchor = "center")

def getteg():
        global E1,\
            L1
        Eget = E1.get()
        for i in range(2, 10):
            Cell = Sheeter.cell(row=i,
                                column=1)
            if Eget.lower() == Cell.value.lower():
                Bile.set(f"Stock: {Sheeter.cell(row=i,
                                                column=2).value}")
                break

        else:
            tkinter.messagebox.showinfo("openpyxl.Error",
                                        "Sorry, this sweet doesn't exist.")

# Button
B1 = Button(root,
            text="Submit", width = 10,
            style="default.TButton",
            command=lambda: getteg())
B1.place(relx = 0.5,
         rely = 0.5,
         anchor = "center")

# Mainloop
root.mainloop()