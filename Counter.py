# Import Statements
from ttkbootstrap import *
from tkinter.messagebox import showinfo
import openpyxl
import sqlite3
import pygame
import segno

# Initializing Pygame
pygame.init()

# Opening Excel
customers = openpyxl.load_workbook("Customers.xlsx")
inventory = openpyxl.load_workbook("Inventory.xlsx")

# Opening the database and cursor
Time = sqlite3.connect("CustNum.db")
Timmy = Time.cursor()

# Window
root = Window(themename = 'vapor')
root.geometry('800x500')
root.title("Billing")

# The Icon Photo
Room = PhotoImage(file="Bill.png") # .gif or .ppm
root.iconphoto(False, Room)

# Variables
ListOfOrders = [] # Bill
SweetNames = [] # Sweets
Sweets = [] # Prices
Quant = [] # Quantity
SalesTax = 1.2 # Tax
Sweet = "" # The Sweet of the Button
SweetPrices = {"Gulab Jamun": 13.25,
               "Rasgulla": 16,
               "Ras Malai": 22,
               "Kaju Katli": 13,
               "Laddu": 15,
               "Sandesh": 20,
               "Jalebi": 30,
               "Lengcha": 17} # Costs of each sweet
Bill = StringVar(value="") # The bill

# Style
style = Style()
style.configure('secondary.TButton',
                font = ("Comic Sans MS",
                        8)) # Blue button
style1 = Style()
style1.configure('warning.TButton',
                 font = ("Comic Sans MS",
                         15)) # Orange Button

# Playing a sound
def Play_Sound(text: str):
    pygame.mixer.Sound(text).play()

# Functions
def GJ():
    global Sweet
    Sweet = "Gulab Jamun"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Gulab Jamun.mp3")

def RG():
    global Sweet
    Sweet = "Rasgulla"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Rasgulla.mp3")
def LC():
    global Sweet
    Sweet = "Lengcha"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Lengcha.mp3")
def KK():
    global Sweet
    Sweet = "Kaju Katli"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Kaju Katli.mp3")
def RaM():
    global Sweet
    Sweet = "Ras Malai"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Ras Malai.mp3")
def SN():
    global Sweet
    Sweet = "Sandesh"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Sandesh.mp3")
def LD():
    global Sweet
    Sweet = "Laddu"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Laddu.mp3")
def JL():
    global Sweet
    Sweet = "Jalebi"

    # Checking the phone no.
    if (len(E3.get()) > 10 or
            len(E3.get()) < 10):
        showinfo("Invalid Phone Number",
                 f"Your Phone Number has {len(E3.get())} digits,"
                 f" which is invalid. Please Try Again.")
        return

    # Play the voice
    Play_Sound("Jalebi.mp3")

#Sweets LabelFrame
SL1 = Labelframe(root,
                 text='Sweets!',
                 bootstyle=INFO)
SL1B1 = Button(SL1,
               text = "Gulab Jamun",
               style='secondary.TButton',
               command=GJ)
SL1B1.pack(padx = 10,
           pady = 10,
           side = 'bottom') # GJ
SL1B2 = Button(SL1,
               text = "Rasgulla",
               style='secondary.TButton', command=RG)
SL1B2.pack(padx = 10,
           pady = 10,
           side = 'bottom') # RG
SL1B3 = Button(SL1,
               text = "Lengcha",
               style='secondary.TButton',
               command=LC)
SL1B3.pack(padx = 10,
           pady = 10,
           side = 'bottom') # LC
SL1B4 = Button(SL1,
               text = "Ras Malai",
               style='secondary.TButton',
               command=RaM)
SL1B4.pack(padx = 10,
           pady = 10,
           side = 'bottom') # RaM
SL1B5 = Button(SL1,
               text = "Kaju Katli",
               style='secondary.TButton',
               command=KK)
SL1B5.pack(padx = 10,
           pady = 10,
           side = 'bottom') # KK
SL1B6 = Button(SL1,
               text = "Laddu",
               style='secondary.TButton',
               command=LD)
SL1B6.pack(padx = 10,
           pady = 10,
           side = 'bottom') # LD
SL1B7 = Button(SL1,
               text = "Jalebi",
               style='secondary.TButton',
               command=JL)
SL1B7.pack(padx = 10,
           pady = 10,
           side = 'bottom') # JL
SL1B8 = Button(SL1,
               text = "Sandesh",
               style='secondary.TButton',
               command=SN)
SL1B8.pack(padx = 10,
           pady = 10,
           side = 'bottom') # SN
SL1.place(anchor = CENTER,
          relx = 0.1,
          rely = 0.475)

# Quantity
QL2 = Labelframe(root,
                 text = "Quantity",
                 bootstyle = SUCCESS) # Labelframe
S1 = Spinbox(QL2,
             from_=1,
             to=200,
             bootstyle=SUCCESS,
             width=7,
             font=("Comic Sans MS",
                   13)) # Spinbox
S1.pack(padx=10,
        pady=10)
QL2.place(relx = 0.89,
          rely = 0.11,
          anchor = 'center')

# Label for order
L1 = Label(root,
           font=("Comic Sans MS",
                 8))
L1.place(relx = 0.485,
         rely = 0.475,
         anchor = 'center')

# Label for bill
L2 = Label(root,
           textvariable=Bill,
           font=("Comic Sans MS", 12),
           justify=CENTER,
           bootstyle = SUCCESS)
L2.place(relx = 0.5,
         rely = 0.775,
         anchor = 'center')

# Adding a sweet
def AddToList():
    global Sweets, \
        Quant, \
        SweetNames, \
        SweetPrices, \
        Sweet, \
        S1, \
        ListOfOrders, \
        Orders # Globalize

    # Finding out stock
    for i in range(2, 9):
       Sheet = inventory.active
       if Sheet.cell(row=i,
                     column=1).value == Sweet:
           if int(Sheet.cell(row=i,
                             column=2).value) <= int(S1.get()):
               showinfo("Out of Stock", "Sweet is out of stock."
                                        " Try again.")
               return

    # Adding the item to the bill
    SweetNames.append(Sweet)
    Sweets.append(SweetPrices[Sweet])
    Quant.append(int(S1.get()))
    ListOfOrders.append(f"Item: {Sweet} |"
                        f" Quant. : {int(S1.get())} |"
                        f" Price: {SweetPrices[Sweet]} |"
                        f" Total: {SweetPrices[Sweet] * int(S1.get())}")
    L1.configure(text="\n".join(ListOfOrders),
                 justify=CENTER)


# Deleting an item
def DeleteFromList():
    global Sweets,\
        Quant,\
        SweetNames, \
        SweetPrices,\
        Sweet,\
        S1 # Globalize

    # Delete
    SweetNames.remove(Sweet)
    Sweets.remove(SweetPrices[Sweet])
    Quant.remove(int(S1.get()))
    ListOfOrders.remove(f"Item: {Sweet} | "
                        f"Quant. : {int(S1.get())} |"
                        f" Price: {SweetPrices[Sweet]} |"
                        f" Total: {SweetPrices[Sweet] * int(S1.get())}")
    L1.configure(text="\n".join(ListOfOrders),
                 justify=CENTER)

# Getting the bill
def GetBill():
     global Sweets,\
         Quant,\
         SweetNames,\
         SweetPrices,\
         Sweet,\
         S1,\
         E2,\
         E3 # Globalize

     # Getting the bill (Configure label)
     NewSweets = [Sweets[i] * Quant[i] for i in range(len(Sweets))]
     Total = sum(NewSweets)
     PreTotal = Total * 0.9
     NewTotal = PreTotal * SalesTax
     Bill.set(f"Total: {Total}\n"
              f"Discount: 10%\n"
              f"Sales Tax: 20%\n"
              f"Grand Total: {int(NewTotal)}")

     # Changing the excel sheets and databases
     Timmy.execute("SELECT * from CustNum")
     for SweetName in SweetNames:
        Sheet = inventory.active
        for i in range(2, 9):
            if Sheet.cell(row=i,
                          column=1).value == SweetName:
                Sheet.cell(row=i,
                           column=2).value -= Quant[SweetNames.index(SweetName)]
     Timmy.execute("SELECT * from CustNum")
     TimeF = Timmy.fetchall()[0][0]
     Timmy.execute("UPDATE CustNum SET Num=:a",
                   {"a": TimeF + 1})
     Sheet = customers.active
     Sheet.cell(row=TimeF + 1,
                column=1).value = E2.get()
     Sheet.cell(row=TimeF + 1,
                column=2).value = E3.get()
     Sheet.cell(row=TimeF + 1,
                column=3).value = NewTotal

     # Creating a QR code
     Q4R = segno.make_qr(f"{L1["text"]}\n"
                         f"Grand Total: {NewTotal}",
                         version=20)  # Create it
     Q4R.save(f"QR.png")  # Save it

     # Pre defined constants
     screen = pygame.display.set_mode((500,
                                       500))
     Clock = pygame.time.Clock()
     running = True

     # Database
     Timmy.execute("INSERT INTO Things_Bought VALUES (?, ?, ?, ?)",
                   [E2.get(),
                    int(E3.get()),
                    L1['text'],
                    NewTotal])

     # QR Code
     image = pygame.transform.scale(pygame.image.load("QR.png"),
                                    (225,
                                     225))
     imger = image.get_rect(center=(250,
                                    250))

     # Pygame
     while running:

         # Event loop
         for event in pygame.event.get():
             if event.type == pygame.QUIT:
                 running = False
             if pygame.key.get_pressed()[pygame.K_ESCAPE]:
                 running = False

         screen.fill("white")
         screen.blit(image,
                     imger)
         pygame.display.update()

# Starting a new bill
def BillToZero():
    global Sweets,\
        Quant,\
        SweetNames,\
        SweetPrices,\
        Sweet,\
        S1,\
        E2,\
        E3 # Globalize

    # Setting all values to None
    ListOfOrders.clear()
    SweetNames.clear()
    Sweets.clear()
    Quant.clear()
    L1.configure(text="")
    L2.configure(text="")
    Bill.set("")
    E2.delete(0,
              END)
    E3.delete(0,
              END)

# Add to Bill
A1 = Button(root,
            text="Add",
            style='warning.TButton',
            width = 7,
            command=AddToList)
A1.place(relx = 0.89,
         rely = 0.293,
         anchor = 'center')

# Delete from bill
D1 = Button(root,
            text="Delete",
            style='warning.TButton',
            width = 7,
            command=DeleteFromList)
D1.place(relx = 0.89,
         rely = 0.455,
         anchor = 'center')

# Get the bill
G1 = Button(root,
            text="Get Bill",
            style='warning.TButton',
            width = 7,
            command = lambda: GetBill())
G1.place(relx = 0.89,
         rely = 0.615,
         anchor = 'center')

# Button
B1 = Button(root,
            text="New Bill",
            style='warning.TButton',
            width = 7,
            command = BillToZero)
B1.place(relx = 0.89,
         rely = 0.825,
         anchor = 'center')

# The customer details section
F1 = Frame(root)
F2 = Frame(F1) # Name
F3 = Frame(F1) # Phone no.
E2 = Entry(F2,
           font=("Comic Sans MS",
                 12),
           width = 27) # Name
E3 = Entry(F3,
           font=("Comic Sans MS",
                 12),
           width = 27) # Phone
L3 = Label(F2,
           font=("Comic Sans MS",
                 12),
           text="Name: ") # Label
L4 = Label(F3,
           font=("Comic Sans MS",
                 12),
           text="Phone No.: ") # Label

# Pack Pack!
L3.pack(padx = 20,
        side=LEFT)
L4.pack(side=LEFT)
E2.pack(padx = 20,
        side=RIGHT)
E3.pack(padx = 20,
        side=RIGHT)
F2.pack(pady = 10)
F3.pack()
F1.place(relx = 0.489,
         rely = 0.17, anchor=CENTER)

root.mainloop()

# Saving the files
Time.commit()
inventory.save("Inventory.xlsx")
customers.save("Customers.xlsx")